"""
Nifty 500 Weekly/Monthly RSI Screener
--------------------------------------
Scans every stock in nifty500.csv (column 'Symbol'), computes Weekly RSI and
Monthly RSI (Wilder's method), and flags stocks where:

    Weekly RSI  < 40   (oversold on the weekly timeframe)
    Monthly RSI > 40   (not oversold on the higher timeframe -> pullback in an
                        otherwise healthy trend, not a full breakdown)

Outputs:
    docs/nifty500_rsi_report.xlsx  - Stock, Current Price, Weekly RSI,
                                      Monthly RSI, % Down from 52-Week High
    docs/index.html                - HTML report with a download link
                                      (served via GitHub Pages)

Alerts:
    Telegram message with the match count + top matches
    Generic webhook POST with the full JSON payload

Environment variables required (set as GitHub Actions secrets):
    TELEGRAM_BOT_TOKEN
    TELEGRAM_CHAT_ID
    WEBHOOK_URL           (optional - skipped if not set)
    PAGES_BASE_URL        (optional - used to build the download link in the
                            HTML report, e.g. https://<user>.github.io/<repo>)
"""

import os
import time
import json
from datetime import datetime, timezone, timedelta

import pandas as pd
import yfinance as yf
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------
CSV_FILE = "nifty500.csv"
OUTPUT_DIR = "docs"
XLSX_NAME = "nifty500_rsi_report.xlsx"
HTML_NAME = "index.html"

RSI_PERIOD = 14
WEEKLY_RSI_MAX = 40      # Weekly RSI < 40
MONTHLY_RSI_MIN = 40     # Monthly RSI > 40

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
WEBHOOK_URL = os.environ.get("WEBHOOK_URL", "")
PAGES_BASE_URL = os.environ.get("PAGES_BASE_URL", "")

IST = timezone(timedelta(hours=5, minutes=30))


# --------------------------------------------------------------------------
# RSI (Wilder's exact method)
# --------------------------------------------------------------------------
def calculate_rsi(close: pd.Series, period: int = RSI_PERIOD) -> pd.Series:
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = -delta.where(delta < 0, 0.0)

    avg_gain = gain.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi


# --------------------------------------------------------------------------
# Data
# --------------------------------------------------------------------------
def load_symbols(csv_file: str) -> list[str]:
    df = pd.read_csv(csv_file)
    if "Symbol" not in df.columns:
        raise ValueError("CSV must contain a 'Symbol' column as the first row header")
    symbols = df["Symbol"].dropna().astype(str).str.strip().tolist()
    return symbols


def fetch_and_score(symbol: str) -> dict | None:
    """Fetch full daily history for one symbol, derive weekly & monthly RSI,
    current price and % down from the 52-week high."""
    ticker = f"{symbol}.NS"
    try:
        hist = yf.Ticker(ticker).history(period="max", interval="1d", auto_adjust=False)
    except Exception:
        return None

    if hist is None or hist.empty or len(hist) < RSI_PERIOD * 5:
        return None

    hist = hist.dropna(subset=["Close"])
    hist.index = pd.to_datetime(hist.index)

    weekly_close = hist["Close"].resample("W-FRI").last().dropna()
    monthly_close = hist["Close"].resample("ME").last().dropna()

    if len(weekly_close) < RSI_PERIOD + 1 or len(monthly_close) < RSI_PERIOD + 1:
        return None

    weekly_rsi = calculate_rsi(weekly_close).iloc[-1]
    monthly_rsi = calculate_rsi(monthly_close).iloc[-1]

    if pd.isna(weekly_rsi) or pd.isna(monthly_rsi):
        return None

    current_price = float(hist["Close"].iloc[-1])
    last_date = hist.index[-1]
    window_52w = hist.loc[hist.index >= last_date - pd.Timedelta(days=365)]
    high_52w = float(window_52w["Close"].max())
    down_from_52w_high_pct = (current_price - high_52w) / high_52w * 100

    return {
        "symbol": symbol,
        "current_price": round(current_price, 2),
        "weekly_rsi": round(float(weekly_rsi), 2),
        "monthly_rsi": round(float(monthly_rsi), 2),
        "down_from_52w_high_pct": round(down_from_52w_high_pct, 2),
    }


def scan_all(symbols: list[str]) -> list[dict]:
    matches = []
    for i, symbol in enumerate(symbols, 1):
        result = fetch_and_score(symbol)
        if result and result["weekly_rsi"] < WEEKLY_RSI_MAX and result["monthly_rsi"] > MONTHLY_RSI_MIN:
            matches.append(result)
        if i % 25 == 0:
            print(f"Scanned {i}/{len(symbols)} symbols...")
        time.sleep(0.2)  # be polite to the data source
    matches.sort(key=lambda r: r["weekly_rsi"])
    return matches


# --------------------------------------------------------------------------
# XLSX report
# --------------------------------------------------------------------------
def write_xlsx(matches: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RSI Screener"

    headers = ["Stock", "Current Price", "Weekly RSI", "Monthly RSI", "% Down from 52W High"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in matches:
        ws.append([
            row["symbol"],
            row["current_price"],
            row["weekly_rsi"],
            row["monthly_rsi"],
            row["down_from_52w_high_pct"],
        ])

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 4)

    ws.freeze_panes = "A2"
    wb.save(path)


# --------------------------------------------------------------------------
# HTML report
# --------------------------------------------------------------------------
def write_html(matches: list[dict], path: str, xlsx_filename: str, run_time_ist: str) -> None:
    download_href = xlsx_filename
    if PAGES_BASE_URL:
        download_href = f"{PAGES_BASE_URL.rstrip('/')}/{xlsx_filename}"

    rows_html = "\n".join(
        f"<tr><td>{r['symbol']}</td><td>{r['current_price']}</td>"
        f"<td>{r['weekly_rsi']}</td><td>{r['monthly_rsi']}</td>"
        f"<td>{r['down_from_52w_high_pct']}%</td></tr>"
        for r in matches
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Nifty 500 Weekly/Monthly RSI Screener</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; background:#0f1115; color:#e6e6e6; margin:0; padding:32px; font-size:18px; }}
  h1 {{ font-size: 2rem; margin-bottom:6px; }}
  .meta {{ color:#9aa0a6; font-size:1.05rem; margin-bottom:22px; }}
  .download {{ display:inline-block; margin-bottom:26px; padding:14px 22px; background:#2563eb; color:#fff; text-decoration:none; border-radius:6px; font-weight:600; font-size:1.1rem; }}
  table {{ border-collapse: collapse; width:100%; background:#171a21; font-size:1.05rem; }}
  th, td {{ padding:14px 16px; text-align:left; border-bottom:1px solid #2a2e37; }}
  th {{ background:#1f4e78; color:#fff; position:sticky; top:0; font-size:1.05rem; }}
  tr:hover {{ background:#20242c; }}
  .empty {{ color:#9aa0a6; padding:16px 0; font-size:1.1rem; }}
</style>
</head>
<body>
  <h1>Nifty 500 &mdash; Weekly RSI &lt; 30 &amp; Monthly RSI &gt; 40</h1>
  <div class="meta">Run: {run_time_ist} IST &nbsp;|&nbsp; Matches: {len(matches)}</div>
  <a class="download" href="{download_href}">Download Excel report</a>
  {"<table><thead><tr><th>Stock</th><th>Current Price</th><th>Weekly RSI</th><th>Monthly RSI</th><th>% Down from 52W High</th></tr></thead><tbody>" + rows_html + "</tbody></table>" if matches else '<div class="empty">No stocks matched the criteria this run.</div>'}
</body>
</html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


# --------------------------------------------------------------------------
# Alerts
# --------------------------------------------------------------------------
def send_telegram(matches: list[dict], report_url: str) -> None:
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("Telegram not configured, skipping.")
        return

    if matches:
        header = f"<b>Nifty 500 RSI Screener</b> \u2014 {len(matches)} match(es)\n"
        header += "Weekly RSI &lt; 30, Monthly RSI &gt; 40\n\n"

        # Fixed-width columns rendered inside <pre> so Telegram shows each
        # stock on exactly one line, monospaced, without wrapping.
        col_symbol, col_num = 11, 7
        table_lines = [
            f"{'Stock':<{col_symbol}}{'CMP':>{col_num}}{'WkRSI':>{col_num}}{'MoRSI':>{col_num}}{'52WH%':>{col_num}}"
        ]
        for r in matches[:25]:
            table_lines.append(
                f"{r['symbol']:<{col_symbol}}"
                f"{r['current_price']:>{col_num}}"
                f"{r['weekly_rsi']:>{col_num}}"
                f"{r['monthly_rsi']:>{col_num}}"
                f"{r['down_from_52w_high_pct']:>{col_num}}"
            )
        table_html = "<pre>" + "\n".join(table_lines) + "</pre>"

        footer = ""
        if len(matches) > 25:
            footer += f"\n...and {len(matches) - 25} more in the report."
        if report_url:
            footer += f'\n<a href="{report_url}">Full report</a>'

        text = header + table_html + footer
    else:
        text = "<b>Nifty 500 RSI Screener</b> \u2014 no stocks matched this run."

    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"},
            timeout=15,
        )
    except Exception as e:
        print(f"Telegram send failed: {e}")


def send_webhook(matches: list[dict], run_time_ist: str) -> None:
    if not WEBHOOK_URL:
        print("Webhook not configured, skipping.")
        return
    payload = {"run_time_ist": run_time_ist, "match_count": len(matches), "matches": matches}
    try:
        requests.post(WEBHOOK_URL, json=payload, timeout=15)
    except Exception as e:
        print(f"Webhook send failed: {e}")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------
def main() -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    run_time_ist = datetime.now(IST).strftime("%Y-%m-%d %H:%M")

    symbols = load_symbols(CSV_FILE)
    print(f"Loaded {len(symbols)} symbols from {CSV_FILE}")

    matches = scan_all(symbols)
    print(f"Found {len(matches)} matching stocks")

    xlsx_path = os.path.join(OUTPUT_DIR, XLSX_NAME)
    html_path = os.path.join(OUTPUT_DIR, HTML_NAME)

    write_xlsx(matches, xlsx_path)
    write_html(matches, html_path, XLSX_NAME, run_time_ist)

    report_url = f"{PAGES_BASE_URL.rstrip('/')}/" if PAGES_BASE_URL else ""
    send_telegram(matches, report_url)
    send_webhook(matches, run_time_ist)

    print("Done.")


if __name__ == "__main__":
    main()
